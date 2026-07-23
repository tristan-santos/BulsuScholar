import csv
import io
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from .supabase_ops import supabase_select
except ImportError:  # pragma: no cover - supports direct backend execution
    from supabase_ops import supabase_select


STUDENT_REPORT_HEADERS = [
    "Student ID",
    "Full Name",
    "Course",
    "Year Level",
    "GWA",
    "Grantor",
    "Record Status",
]


def _text(value: Any, fallback: str = "-") -> str:
    rendered = str(value or "").strip()
    return rendered or fallback


def _student_name(data: dict[str, Any]) -> str:
    explicit = data.get("fullName") or data.get("fullname") or data.get("name")
    if explicit:
        return _text(explicit)
    return _text(" ".join(filter(None, [data.get("fname"), data.get("mname"), data.get("lname")])))


def _student_current_stage(data: dict[str, Any]) -> str:
    direct = (
        data.get("currentStepLabel")
        or data.get("currentStage")
        or data.get("trackingStage")
        or data.get("applicationStage")
    )
    if direct:
        return _text(direct)

    scholarships = data.get("scholarships") if isinstance(data.get("scholarships"), list) else []
    for scholarship in scholarships:
        if not isinstance(scholarship, dict):
            continue
        status = _text(scholarship.get("status"), "").lower()
        if any(value in status for value in ("rejected", "cancelled", "withdrawn")):
            continue
        tracking = scholarship.get("tracking") if isinstance(scholarship.get("tracking"), dict) else {}
        stage = (
            scholarship.get("currentStepLabel")
            or scholarship.get("currentStage")
            or scholarship.get("stage")
            or tracking.get("currentStepLabel")
            or tracking.get("currentStage")
        )
        if stage:
            return _text(stage)
    return "Account Created"


def _student_grantor(data: dict[str, Any]) -> str:
    direct = (
        data.get("grantor")
        or data.get("grantorName")
        or data.get("providerName")
        or data.get("providerType")
    )
    return _text(direct, "N/A")


def _student_id(value: Any) -> str:
    return _text(value).removeprefix("roster_")


def _flatten_student_row(raw: dict[str, Any]) -> dict[str, Any]:
    nested = raw.get("data") if isinstance(raw.get("data"), dict) else {}
    data = {**raw, **nested}
    student_id = data.get("studentnumber") or data.get("studentNumber") or data.get("studentId") or data.get("id") or raw.get("id")
    record_status = data.get("recordStatus") or data.get("status")
    if not record_status:
        record_status = "Archived" if data.get("archived") is True else "Active"
    return {
        "id": _student_id(student_id),
        "fullName": _student_name(data),
        "email": _text(data.get("email")),
        "cpNumber": _text(data.get("cpNumber") or data.get("contactNumber") or data.get("number")),
        "course": _text(data.get("course")),
        "yearLevel": _text(data.get("year") or data.get("yearLevel")),
        "gwa": _text(data.get("gwa") or data.get("currentGwa") or data.get("currentGWA")),
        "grantor": _student_grantor(data),
        "recordStatus": _text(record_status),
    }


def _student_report_filters(payload: dict[str, Any]) -> dict[str, str]:
    filters = payload.get("filters") if isinstance(payload.get("filters"), dict) else payload
    return {
        "search": _text(filters.get("search"), "").lower(),
        "course": _text(filters.get("course"), "All"),
        "year": _text(filters.get("year"), "All"),
        "view": _text(filters.get("view"), "students").lower(),
    }


def build_student_report(payload: dict[str, Any]) -> dict[str, Any]:
    filters = _student_report_filters(payload)
    payload_rows = payload.get("rows") if isinstance(payload.get("rows"), list) else []
    if payload_rows:
        students = [_flatten_student_row(row) for row in payload_rows if isinstance(row, dict)]
        filtered = students
    else:
        result = supabase_select("students", limit=10000)
        if not result.get("ok"):
            raise RuntimeError(result.get("reason") or result.get("detail") or "Unable to load students.")
        students = [_flatten_student_row(row) for row in result.get("rows", [])]

        def include(student: dict[str, Any]) -> bool:
            if filters["view"] == "archived" and student["recordStatus"] != "Archived":
                return False
            if filters["view"] == "students" and student["recordStatus"] == "Archived":
                return False
            if filters["course"] != "All" and student["course"] != filters["course"]:
                return False
            if filters["year"] != "All" and student["yearLevel"] != filters["year"]:
                return False
            haystack = f"{student['id']} {student['fullName']} {student['email']}".lower()
            return not filters["search"] or filters["search"] in haystack

        filtered = sorted((student for student in students if include(student)), key=lambda item: item["fullName"].lower())
    rows = [[student[key] for key in ("id", "fullName", "course", "yearLevel", "gwa", "grantor", "recordStatus")] for student in filtered]
    filter_label = f"View: {filters['view']} | Search: {filters['search'] or '-'} | Course: {filters['course']} | Year: {filters['year']}"
    return {
        "key": "students",
        "title": "Student Management Report",
        "description": "Backend-generated student records using the current management filters.",
        "filterLabel": filter_label,
        "columns": STUDENT_REPORT_HEADERS,
        "rows": rows,
        "stats": [
            {"label": "Records", "value": len(filtered)},
            {"label": "Active", "value": sum(item["recordStatus"] == "Active" for item in filtered)},
            {"label": "Archived", "value": sum(item["recordStatus"] == "Archived" for item in filtered)},
        ],
    }


def build_student_report_pdf_bytes(payload: dict[str, Any]) -> tuple[bytes, str]:
    report = build_student_report(payload)
    report["subtitle"] = "Student Management"
    report["headers"] = report["columns"]
    filename = f"student-management-{datetime.now().strftime('%Y%m%d-%H%M%S')}.pdf"
    return build_report_pdf_bytes(report), filename


def build_student_report_excel_bytes(payload: dict[str, Any]) -> tuple[bytes, str]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as error:  # pragma: no cover - dependency guard
        raise RuntimeError("openpyxl is required for Excel report generation.") from error

    report = build_student_report(payload)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students"
    sheet.append([report["title"]])
    sheet.append([report["filterLabel"]])
    sheet.append([])
    sheet.append(report["columns"])
    for row in report["rows"]:
        sheet.append(row)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(report["columns"]))
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(report["columns"]))
    sheet["A1"].font = Font(size=16, bold=True, color="00633C")
    sheet["A2"].font = Font(size=10, color="526176")
    for cell in sheet[4]:
        cell.fill = PatternFill("solid", fgColor="00633C")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:{get_column_letter(len(report['columns']))}{max(4, sheet.max_row)}"
    for column_index, column_cells in enumerate(sheet.columns, start=1):
        width = min(42, max(12, max(len(str(cell.value or "")) for cell in column_cells) + 2))
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    for row in sheet.iter_rows(min_row=5):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = io.BytesIO()
    workbook.save(output)
    filename = f"student-management-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    return output.getvalue(), filename


def build_csv_bytes(headers: list[str], rows: list[list[Any]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8")


def _report_column_widths(headers: list[str], available_width: float) -> list[float]:
    if headers == STUDENT_REPORT_HEADERS:
        weights = [0.12, 0.18, 0.24, 0.09, 0.08, 0.14, 0.15]
        return [available_width * weight for weight in weights]
    if not headers:
        return []
    return [available_width / len(headers)] * len(headers)


def build_report_pdf_bytes(payload: dict[str, Any]) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        from pypdf import PdfReader, PdfWriter
    except ImportError as error:  # pragma: no cover - dependency guard
        raise RuntimeError("reportlab and pypdf are required for Python PDF report generation.") from error

    font_regular = "Times-Roman"
    font_bold = "Times-Bold"
    windows_fonts = Path("C:/Windows/Fonts")
    times_regular = windows_fonts / "times.ttf"
    times_bold = windows_fonts / "timesbd.ttf"
    if times_regular.exists() and times_bold.exists():
        pdfmetrics.registerFont(TTFont("TimesNewRoman", str(times_regular)))
        pdfmetrics.registerFont(TTFont("TimesNewRoman-Bold", str(times_bold)))
        font_regular = "TimesNewRoman"
        font_bold = "TimesNewRoman-Bold"

    overlay_buffer = io.BytesIO()
    template_path = Path(__file__).resolve().parents[1] / "public" / "Templates" / "FORMATTED_REPORT.pdf"
    template_reader = PdfReader(str(template_path)) if template_path.exists() else None
    template_page = template_reader.pages[0] if template_reader else None
    page_width = float(template_page.mediabox.width) if template_page else letter[0]
    page_height = float(template_page.mediabox.height) if template_page else letter[1]
    page_size = (page_width, page_height)
    margin_left = 36
    margin_right = 36
    margin_top = 190 if template_page else 72
    margin_bottom = 178 if template_page else 62
    doc = SimpleDocTemplate(
        overlay_buffer,
        pagesize=page_size,
        leftMargin=margin_left,
        rightMargin=margin_right,
        topMargin=margin_top,
        bottomMargin=margin_bottom,
    )
    styles = getSampleStyleSheet()
    for style_name in ("Title", "Heading2", "BodyText"):
        styles[style_name].fontName = font_regular
        styles[style_name].fontSize = 10
        styles[style_name].leading = 12
    styles["Title"].fontName = font_bold
    styles["Heading2"].fontName = font_bold
    styles.add(styles["BodyText"].clone("ReportTableHeader"))
    styles.add(styles["BodyText"].clone("ReportTableCell"))
    styles["ReportTableHeader"].fontName = font_bold
    styles["ReportTableHeader"].fontSize = 10
    styles["ReportTableHeader"].leading = 12
    styles["ReportTableHeader"].textColor = colors.white
    styles["ReportTableCell"].fontName = font_regular
    styles["ReportTableCell"].fontSize = 10
    styles["ReportTableCell"].leading = 12
    styles["ReportTableCell"].wordWrap = "CJK"
    story = [
        Paragraph("BulsuScholar", styles["Title"]),
        Paragraph(payload.get("title") or "Report", styles["Heading2"]),
        Paragraph(payload.get("subtitle") or "", styles["BodyText"]),
        Spacer(1, 10),
        Paragraph(f"Generated: {datetime.now().strftime('%b %d, %Y %I:%M %p')}", styles["BodyText"]),
    ]
    if payload.get("filterLabel"):
        story.append(Paragraph(f"Filters: {payload['filterLabel']}", styles["BodyText"]))
    story.append(Spacer(1, 12))

    raw_headers = payload.get("headers") or payload.get("columns") or []
    headers = [
        item.get("label", "")
        if isinstance(item, dict)
        else str(item)
        for item in raw_headers
    ]
    rows = payload.get("rows") or []
    table_data = []
    if headers:
        table_data.append([Paragraph(header, styles["ReportTableHeader"]) for header in headers])
    for row in rows:
        table_data.append([Paragraph(_text(value), styles["ReportTableCell"]) for value in row])
    if table_data:
        column_widths = _report_column_widths(headers, page_width - margin_left - margin_right)
        table = Table(table_data, colWidths=column_widths, repeatRows=1, hAlign="LEFT", splitByRow=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#00633c")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#b7c8be")),
            ("FONTNAME", (0, 0), (-1, 0), font_bold),
            ("FONTNAME", (0, 1), (-1, -1), font_regular),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(table)
    else:
        story.append(Paragraph("No rows available for the selected report.", styles["BodyText"]))

    def draw_template_marker_masks(canvas, _document):
        if not template_page:
            return
        canvas.saveState()
        canvas.setFillColor(colors.white)
        canvas.setStrokeColor(colors.white)
        canvas.rect(0, 175, page_width, page_height - 370, fill=1, stroke=0)
        canvas.rect(66, 820, 320, 30, fill=1, stroke=0)
        canvas.rect(66, 208, 420, 30, fill=1, stroke=0)
        canvas.restoreState()

    doc.build(story, onFirstPage=draw_template_marker_masks, onLaterPages=draw_template_marker_masks)
    overlay_buffer.seek(0)
    if not template_page:
        return overlay_buffer.getvalue()

    overlay_reader = PdfReader(overlay_buffer)
    writer = PdfWriter()
    for overlay_page in overlay_reader.pages:
        page = template_page.clone(writer)
        page.merge_page(overlay_page)
        writer.add_page(page)

    output = io.BytesIO()
    writer.write(output)
    return output.getvalue()
